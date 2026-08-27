import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

ROOT = '/home/user/kirk-ogren-game'
idx = json.load(open(f'{ROOT}/data/categories.json'))
order = [c['id'] for c in idx['categories']]
titles = {c['id']: c['title'] for c in idx['categories']}

rows = []
for cid in order:
    doc = json.load(open(f'{ROOT}/data/facts/{cid}.json'))
    for f in doc['facts']:
        if f['kind'] == 'boolean':
            stated = 'TRUE' if f['answer'] else 'FALSE'
        else:
            stated = f['value']
        rows.append({
            'claim': f['claim'],
            'stated': stated,
            'context': f.get('context', ''),
            'deeper': f.get('goDeeper', ''),
            'source': f['source']['url'],
            'source_title': f['source']['title'],
            'cat': titles[cid],
            'id': f['id'],
            'kind': f['kind'],
            'unit': f.get('unit', ''),
            'approx': 'approx' if f.get('approx') else 'exact',
            'sensitive': 'yes' if f.get('sensitive') else '',
        })

ARIAL = 'Arial'
HDR_FILL = PatternFill('solid', fgColor='16203A')
EDIT_FILL = PatternFill('solid', fgColor='FFFF00')
ALT_FILL = PatternFill('solid', fgColor='F2F4F7')
THIN = Side(style='thin', color='C9CFDA')
BORDER = Border(bottom=THIN)

wb = Workbook()

# ---------------- Facts sheet ----------------
ws = wb.active
ws.title = 'Facts'

cols = [
    ('Verdict', 12), ('Claim', 52), ('Stated answer', 15),
    ('Corrected answer', 18), ('Notes', 34),
    ('Context (shown to players)', 60), ('Go deeper (shown to players)', 60),
    ('Source', 46), ('Category', 26), ('Fact ID', 24),
    ('Kind', 10), ('Unit', 16), ('Exact / approx', 14), ('Sensitive', 11),
]
for i, (name, w) in enumerate(cols, start=1):
    c = ws.cell(row=1, column=i, value=name)
    c.font = Font(name=ARIAL, bold=True, color='FFFFFF', size=10)
    c.fill = HDR_FILL
    c.alignment = Alignment(vertical='center', wrap_text=True)
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 30

for r, d in enumerate(rows, start=2):
    vals = ['', d['claim'], d['stated'], '', '', d['context'], d['deeper'],
            d['source_title'], d['cat'], d['id'], d['kind'], d['unit'],
            d['approx'], d['sensitive']]
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=r, column=i, value=v)
        c.font = Font(name=ARIAL, size=10)
        c.alignment = Alignment(vertical='top', wrap_text=(i in (2, 6, 7)))
        c.border = BORDER
        if i in (1, 4, 5):
            c.fill = EDIT_FILL
        elif r % 2 == 0:
            c.fill = ALT_FILL
    link = ws.cell(row=r, column=8)
    link.hyperlink = d['source']
    link.font = Font(name=ARIAL, size=10, color='0F6E64', underline='single')
    ws.cell(row=r, column=3).alignment = Alignment(vertical='top', horizontal='left')

last = len(rows) + 1
dv = DataValidation(type='list', formula1='"OK,Wrong,Unclear"', allow_blank=True,
                    showDropDown=False, promptTitle='Verdict',
                    prompt='OK if the stated answer is right. Wrong if it is not — put the right one in Corrected answer. Unclear if the sources disagree.')
ws.add_data_validation(dv)
dv.add(f'A2:A{last}')

ws.freeze_panes = 'D2'
ws.auto_filter.ref = f'A1:N{last}'

# ---------------- Progress sheet ----------------
s = wb.create_sheet('Progress', 0)
s.column_dimensions['A'].width = 34
s.column_dimensions['B'].width = 12
for col in 'DEFG':
    s.column_dimensions[col].width = 30 if col == 'D' else 12

def put(cell, value, **kw):
    c = s[cell]; c.value = value
    c.font = Font(name=ARIAL, size=kw.get('size', 10), bold=kw.get('bold', False),
                  color=kw.get('color', '000000'))
    if kw.get('wrap'):
        c.alignment = Alignment(wrap_text=True, vertical='top')
    return c

put('A1', 'Give or Take — fact review', bold=True, size=14)
put('A2', 'Every fact below is an unreviewed draft written by Claude. Check the stated answer against the source, then set a verdict.', wrap=True)
s.merge_cells('A2:G2'); s.row_dimensions[2].height = 30

put('A4', 'How to use this', bold=True, size=11)
put('A5', 'Fill in the three yellow columns on the Facts sheet. Everything else is reference.', wrap=True)
s.merge_cells('A5:G5')
put('A6', 'Verdict'); put('B6', 'OK  /  Wrong  /  Unclear — pick from the dropdown')
put('A7', 'Corrected answer'); put('B7', 'Only when the verdict is Wrong. Put the right value here.')
put('A8', 'Notes'); put('B8', 'Anything the wording should say differently, or a better source.')
put('A10', 'Example of a completed row', bold=True, size=11)
for i, (h, v) in enumerate([('Verdict', 'Wrong'), ('Claim', 'Lincoln was assassinated'),
                            ('Stated answer', '1866'), ('Corrected answer', '1865'),
                            ('Notes', 'April 14, 1865. Check the NPS page too.')]):
    put(f'A{11+i}', h, color='5C6580'); put(f'B{11+i}', v, wrap=True)
    s[f'B{11+i}'].fill = EDIT_FILL

put('A17', 'Where the review stands', bold=True, size=11)
labels = [('Facts in total', f'=COUNTA(Facts!J2:J{last})'),
          ('Marked OK', f'=COUNTIF(Facts!A2:A{last},"OK")'),
          ('Marked Wrong', f'=COUNTIF(Facts!A2:A{last},"Wrong")'),
          ('Marked Unclear', f'=COUNTIF(Facts!A2:A{last},"Unclear")'),
          ('Still to check', '=B18-B19-B20-B21')]
for i, (lab, formula) in enumerate(labels):
    put(f'A{18+i}', lab)
    c = put(f'B{18+i}', formula, bold=(i == 4))
    c.alignment = Alignment(horizontal='right')

put('D17', 'By period', bold=True, size=11)
for i, h in enumerate(['Period', 'Facts', 'OK', 'Wrong']):
    c = put(f'{"DEFG"[i]}18', h, bold=True, color='FFFFFF')
    c.fill = HDR_FILL
for i, cid in enumerate(order):
    r = 19 + i
    put(f'D{r}', titles[cid])
    put(f'E{r}', f'=COUNTIF(Facts!$I$2:$I${last},D{r})').alignment = Alignment(horizontal='right')
    put(f'F{r}', f'=COUNTIFS(Facts!$I$2:$I${last},D{r},Facts!$A$2:$A${last},"OK")').alignment = Alignment(horizontal='right')
    put(f'G{r}', f'=COUNTIFS(Facts!$I$2:$I${last},D{r},Facts!$A$2:$A${last},"Wrong")').alignment = Alignment(horizontal='right')

out = f'{ROOT}/review/give-or-take-facts.xlsx'
os.makedirs(f'{ROOT}/review', exist_ok=True)
wb.save(out)
print('wrote', out, 'with', len(rows), 'facts')
